"""
MediaTrack API — FastAPI + PostgreSQL.

Run:
    pip install -r requirements.txt
    export DATABASE_URL=postgresql://mediatrack:mediatrack@localhost:5432/mediatrack
    python seed.py
    uvicorn app:app --reload --port 8000

Configuration is read from the environment — see .env.example.
Docs auto-generated at http://localhost:8000/docs
"""

import logging
import os
import sys
from datetime import date, timedelta
from pathlib import Path

from fastapi import Depends, FastAPI, File, HTTPException, Request, Response, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from auth import hash_pw, make_token, read_token, verify_pw
from database import database_url, get_conn, init_db

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s %(message)s",
    datefmt="%d-%b %H:%M:%S",
)
log = logging.getLogger("mediatrack")

ENV = os.environ.get("ENVIRONMENT", "development").lower()
IS_PROD = ENV == "production"


def _env_bool(name: str, default: bool) -> bool:
    v = os.environ.get(name)
    if v is None:
        return default
    return v.strip().lower() in ("1", "true", "yes", "on")


# Cookies must be Secure over HTTPS in production; plain HTTP locally.
COOKIE_SECURE = _env_bool("COOKIE_SECURE", IS_PROD)

app = FastAPI(title="MediaTrack", version="1.0.0")


@app.exception_handler(Exception)
async def unhandled(request, exc):
    """Never leak a raw traceback to the UI — log it, return clean JSON."""
    from fastapi.responses import JSONResponse
    log.exception("Unhandled error on %s %s", request.method, request.url.path)
    return JSONResponse(status_code=500, content={"detail": "Internal error — check server logs"})

# CORS. The dashboard is served by this same app, so in production the browser
# makes same-origin requests and needs no CORS at all. CORS_ORIGINS exists for
# the cases that do: the file:// demo mode and any separately hosted frontend.
#   CORS_ORIGINS="*"                                  → open (dev default)
#   CORS_ORIGINS="https://mediatrack.onrender.com"    → locked down (production)
_origins = [o.strip() for o in os.environ.get(
    "CORS_ORIGINS", "" if IS_PROD else "*").split(",") if o.strip()]
if _origins == ["*"]:
    # Credentialed requests are impossible with a wildcard (browsers reject it),
    # so advertise the wildcard honestly instead of pretending cookies work.
    app.add_middleware(
        CORSMiddleware,
        allow_origins=["*"],
        allow_methods=["*"],
        allow_headers=["*"],
    )
    if IS_PROD:
        log.warning("CORS_ORIGINS='*' in production — set explicit origins.")
elif _origins:
    app.add_middleware(
        CORSMiddleware,
        allow_origins=_origins,
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )
    log.info("CORS restricted to: %s", ", ".join(_origins))
else:
    log.info("CORS disabled — dashboard is served same-origin.")

init_db()


def ensure_default_admin() -> None:
    with get_conn() as conn:
        if conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]:
            return
        username = os.environ.get("ADMIN_USERNAME", "admin").strip().lower()
        password = os.environ.get("ADMIN_PASSWORD", "")
        if not password:
            if IS_PROD:
                raise RuntimeError(
                    "No users exist and ADMIN_PASSWORD is not set. Set ADMIN_USERNAME / "
                    "ADMIN_PASSWORD so the first admin can be created safely."
                )
            password = "ajanta123"   # local dev only
        salt, ph = hash_pw(password)
        conn.execute(
            "INSERT INTO users (username, name, role, team, salt, pw_hash, created_at)"
            " VALUES (?,?,?,?,?,?,?)",
            (username, os.environ.get("ADMIN_NAME", "Admin"), "admin", "all", salt, ph,
             date.today().isoformat()),
        )
        conn.commit()
        log.warning("Bootstrap admin created — username '%s'. Change the password after "
                    "first login.", username)


ensure_default_admin()


def bootstrap_inventory() -> None:
    """
    First boot against an empty database: load the real screens + client
    directory, exactly like the .exe launcher does. A no-op once screens exist.
    Set SEED_ON_START=false to start with a completely empty database.
    """
    if not _env_bool("SEED_ON_START", True):
        return
    try:
        from seed import seed_inventory
        if seed_inventory():
            log.info("First run: loaded the seed screens and client directory.")
    except Exception:                                        # noqa: BLE001
        log.exception("Inventory seeding skipped")


bootstrap_inventory()


def current_user(request: Request):
    tok = request.cookies.get("mt_session")
    if not tok:
        auth = request.headers.get("authorization", "")
        if auth.lower().startswith("bearer "):
            tok = auth[7:]
    if not tok:
        tok = request.query_params.get("token")
    parsed = read_token(tok) if tok else None
    if not parsed:
        return None
    uid, tver = parsed
    with get_conn() as conn:
        r = conn.execute(
            "SELECT id, username, name, role, team, can_upload, is_active, token_version"
            " FROM users WHERE id = ?", (uid,)
        ).fetchone()
        if not r or r["token_version"] != tver:
            return None   # token from before a logout-all — invalid everywhere
        if not r["is_active"]:
            return None   # disabled while signed in — every session stops here
        return dict(r)


def require_user(request: Request) -> dict:
    u = current_user(request)
    if not u:
        raise HTTPException(401, "Login required")
    return u


def require_admin(user: dict = Depends(require_user)) -> dict:
    if user["role"] != "admin":
        raise HTTPException(403, "Access denied — admin only")
    return user


def require_upload(user: dict = Depends(require_user)) -> dict:
    """Admins always can; sales users need the can_upload permission."""
    if user["role"] != "admin" and not user.get("can_upload"):
        raise HTTPException(
            403, "You don't have media-upload permission — ask your admin to enable it"
        )
    return user


# ---------------------------------------------------------------- teams
#
# Odisha and Raipur are two business units on one installation. Isolation is
# enforced here, in the API layer, on top of a `team` column on every business
# row — not in the browser. Every read filters on the caller's team and every
# lookup by id is scoped the same way, so editing a URL, an id or a filter in
# the frontend cannot reach the other team's data.
#
# 'all' is not a team: it is the admin's both-teams view. It is only ever stored
# on a user row and never on a screen, client, campaign, slot or booking.

TEAMS = ("odisha", "raipur")
ALL_TEAMS = "all"
TEAM_NAMES = {"odisha": "Odisha Team", "raipur": "Raipur Team", ALL_TEAMS: "All Teams"}


def normalize_team(value: str | None) -> str | None:
    """'Odisha' → 'odisha', 'Admin'/'All Teams' → 'all', anything else → None."""
    v = (value or "").strip().lower().replace(" ", "_").replace("-", "_")
    if v in ("all", "all_teams", "admin", "both"):
        return ALL_TEAMS
    return v if v in TEAMS else None


def user_sees_all_teams(user: dict) -> bool:
    """Admins, and any user explicitly assigned the 'all' team, span both."""
    return user.get("role") == "admin" or user.get("team") == ALL_TEAMS


class TeamScope:
    """The team filter for one request: who is asking and which team they are
    looking at. Restricted users can only ever hold their own team here."""

    def __init__(self, user: dict, active: str):
        self.user = user
        self.active = active

    @property
    def all_teams(self) -> bool:
        return self.active == ALL_TEAMS

    @property
    def teams(self) -> list:
        return list(TEAMS) if self.all_teams else [self.active]

    def where(self, column: str = "team") -> tuple:
        """SQL predicate + params for a WHERE clause. Always safe to inline:
        `column` is supplied by this module, never by the request."""
        if self.all_teams:
            return "TRUE", []
        return f"{column} = ?", [self.active]

    def allows(self, team: str | None) -> bool:
        return self.all_teams or team == self.active

    def guard(self, row, what: str = "Record", column: str = "team"):
        """404 (not 403) for another team's row — a team must not even be able
        to learn that an id exists on the other side."""
        if row is None or not self.allows(row[column]):
            raise HTTPException(404, f"{what} not found")
        return row

    def write_team(self, requested: str | None = None) -> str:
        """Which team a newly created record belongs to."""
        if not self.all_teams:
            return self.active            # requested value is ignored on purpose
        team = normalize_team(requested)
        if team in TEAMS:
            return team
        raise HTTPException(
            422, "Choose a team (Odisha or Raipur) — you are viewing All Teams")


def team_scope(request: Request, user: dict = Depends(require_user)) -> TeamScope:
    """The one place a request's team is decided. A team user's scope comes from
    their user row and nowhere else; the X-Team header / ?team= parameter is
    honoured only for users who legitimately span both teams."""
    if user_sees_all_teams(user):
        requested = normalize_team(
            request.headers.get("x-team") or request.query_params.get("team"))
        return TeamScope(user, requested or ALL_TEAMS)
    if user.get("team") not in TEAMS:
        raise HTTPException(403, "No team assigned to your account — ask your admin")
    return TeamScope(user, user["team"])


def require_admin_scope(scope: TeamScope = Depends(team_scope)) -> TeamScope:
    if scope.user["role"] != "admin":
        raise HTTPException(403, "Access denied — admin only")
    return scope


def require_upload_scope(scope: TeamScope = Depends(team_scope)) -> TeamScope:
    if scope.user["role"] != "admin" and not scope.user.get("can_upload"):
        raise HTTPException(
            403, "You don't have media-upload permission — ask your admin to enable it"
        )
    return scope


@app.get("/api/teams")
def list_teams(user: dict = Depends(require_user)):
    """The teams this user may look at, and which one is active."""
    with get_conn() as conn:
        rows = [dict(r) for r in conn.execute(
            "SELECT code, name FROM teams ORDER BY code")]
    if user_sees_all_teams(user):
        return {"teams": rows + [{"code": ALL_TEAMS, "name": TEAM_NAMES[ALL_TEAMS]}],
                "can_switch": True, "user_team": user.get("team") or ALL_TEAMS}
    return {"teams": [r for r in rows if r["code"] == user.get("team")],
            "can_switch": False, "user_team": user.get("team")}


def log_activity(user_name: str, action: str, detail: str = "",
                 team: str = ALL_TEAMS) -> None:
    try:
        with get_conn() as conn:
            conn.execute(
                "INSERT INTO activity_log (user_name, action, detail, team)"
                " VALUES (?,?,?,?)",
                (user_name, action, detail[:300], team),
            )
            conn.commit()
    except Exception:                                    # noqa: BLE001
        log.exception("activity log write failed")       # never break the request


class LoginIn(BaseModel):
    username: str = Field(min_length=1, max_length=60)
    password: str = Field(min_length=1, max_length=120)


@app.post("/api/login")
def login(body: LoginIn, response: Response):
    with get_conn() as conn:
        r = conn.execute(
            "SELECT * FROM users WHERE LOWER(username) = LOWER(?)",
            (body.username.strip(),),
        ).fetchone()
    if not r or not verify_pw(body.password, r["salt"], r["pw_hash"]):
        log.warning("Failed login for '%s'", body.username)
        raise HTTPException(401, "Incorrect username or password.")
    if not r["is_active"]:
        log.warning("Login blocked for disabled account '%s'", body.username)
        raise HTTPException(403, "This account has been disabled. Ask your administrator "
                                 "to re-enable it.")
    token = make_token(r["id"], r["token_version"])
    response.set_cookie(
        "mt_session", token, httponly=True, samesite="lax",
        secure=COOKIE_SECURE, max_age=30 * 86400,
    )
    log.info("Login: %s (%s / %s)", r["name"], r["role"], r["team"])
    log_activity(r["name"], "login", "", r["team"])
    return {"name": r["name"], "role": r["role"], "username": r["username"],
            "team": r["team"], "team_name": TEAM_NAMES.get(r["team"], r["team"]),
            "can_switch_teams": user_sees_all_teams(dict(r)),
            "can_upload": bool(r["can_upload"]) or r["role"] == "admin",
            "token": token}


@app.post("/api/logout")
def logout(response: Response):
    response.delete_cookie("mt_session")
    return {"ok": True}


@app.get("/api/me")
def me(user: dict = Depends(require_user)):
    return {"name": user["name"], "role": user["role"], "username": user["username"],
            "team": user.get("team"),
            "team_name": TEAM_NAMES.get(user.get("team"), user.get("team")),
            "can_switch_teams": user_sees_all_teams(user),
            "can_upload": bool(user.get("can_upload")) or user["role"] == "admin"}


class PasswordChange(BaseModel):
    old_password: str = Field(min_length=1, max_length=120)
    new_password: str = Field(min_length=6, max_length=120)


@app.post("/api/me/password")
def change_password(body: PasswordChange, response: Response,
                    user: dict = Depends(require_user)):
    """Change own password. Also invalidates every other device's session."""
    with get_conn() as conn:
        r = conn.execute("SELECT * FROM users WHERE id = ?", (user["id"],)).fetchone()
        if not verify_pw(body.old_password, r["salt"], r["pw_hash"]):
            raise HTTPException(401, "Current password is wrong")
        salt, ph = hash_pw(body.new_password)
        new_ver = r["token_version"] + 1
        conn.execute(
            "UPDATE users SET salt = ?, pw_hash = ?, token_version = ? WHERE id = ?",
            (salt, ph, new_ver, user["id"]),
        )
        conn.commit()
    token = make_token(user["id"], new_ver)   # fresh token so THIS device stays in
    response.set_cookie("mt_session", token, httponly=True, samesite="lax",
                        secure=COOKIE_SECURE, max_age=30 * 86400)
    log_activity(user["name"], "password_changed", "")
    return {"ok": True, "token": token}


@app.post("/api/logout-all")
def logout_all(response: Response, user: dict = Depends(require_user)):
    """Sign out from every device (invalidates all tokens, everywhere)."""
    with get_conn() as conn:
        conn.execute(
            "UPDATE users SET token_version = token_version + 1 WHERE id = ?",
            (user["id"],),
        )
        conn.commit()
    response.delete_cookie("mt_session")
    log_activity(user["name"], "logout_all_devices", "")
    return {"ok": True}


# ---- user management (admin) ----

class UserIn(BaseModel):
    username: str = Field(min_length=2, max_length=60)
    name: str = Field(min_length=2, max_length=80)
    password: str = Field(min_length=6, max_length=120)
    role: str = Field(pattern="^(admin|sales)$")
    # Odisha / Raipur / Admin (= all teams). Defaults to Odisha so the existing
    # single-team workflow keeps working unchanged.
    team: str = Field(default="odisha", max_length=20)


def _user_team(role: str, team: str | None) -> str:
    """Admins always span both teams; a sales user must sit in exactly one."""
    if role == "admin":
        return ALL_TEAMS
    t = normalize_team(team)
    if t not in TEAMS:
        raise HTTPException(422, "Team must be Odisha or Raipur")
    return t


@app.get("/api/users")
def list_users(scope: TeamScope = Depends(require_admin_scope)):
    """Admins see their own team's users; the All Teams view lists everyone."""
    where, params = scope.where("team")
    with get_conn() as conn:
        rows = conn.execute(
            f"SELECT id, username, name, role, team, can_upload, is_active, created_at"
            f" FROM users WHERE {where} OR team = 'all' ORDER BY id", params)
        return [dict(r) for r in rows]


@app.post("/api/users", status_code=201)
def add_user(u: UserIn, scope: TeamScope = Depends(require_admin_scope)):
    team = _user_team(u.role, u.team)
    # An admin looking at one team can only create users inside that team.
    if not scope.all_teams and team != scope.active:
        raise HTTPException(403, f"You can only add users to the {scope.active} team")
    salt, ph = hash_pw(u.password)
    with get_conn() as conn:
        try:
            cur = conn.execute(
                "INSERT INTO users (username, name, role, team, salt, pw_hash, created_at)"
                " VALUES (?,?,?,?,?,?,?)",
                (u.username.strip().lower(), u.name.strip(), u.role, team, salt, ph,
                 date.today().isoformat()),
            )
        except Exception:
            raise HTTPException(409, "That username already exists")
        conn.commit()
        log.info("User added: %s (%s / %s)", u.username, u.role, team)
        log_activity(scope.user["name"], "user_added",
                     f"{u.name} ({u.role}, {TEAM_NAMES.get(team, team)})", team)
        return {"created": True, "user_id": cur.lastrowid, "team": team}


def _load_target(conn, user_id: int, scope: TeamScope):
    """The user row an admin is acting on, or 404. An admin scoped to one team
    can only reach that team's users (all-team admins are visible but only
    manageable from the All Teams view, where the scope covers them)."""
    row = conn.execute(
        "SELECT id, username, name, role, team, can_upload, is_active"
        " FROM users WHERE id = ?", (user_id,)).fetchone()
    if not row or not (scope.all_teams or row["team"] == scope.active):
        raise HTTPException(404, "User not found")
    return row


def _active_admin_count(conn, exclude_id: int | None = None) -> int:
    sql = "SELECT COUNT(*) FROM users WHERE role = 'admin' AND is_active = TRUE"
    params = []
    if exclude_id is not None:
        sql += " AND id <> ?"
        params.append(exclude_id)
    return conn.execute(sql, params).fetchone()[0]


class UserTeamIn(BaseModel):
    team: str = Field(max_length=20)


@app.patch("/api/users/{user_id}/team")
def set_user_team(user_id: int, body: UserTeamIn,
                  scope: TeamScope = Depends(require_admin_scope)):
    """Move a user to the other team (or to all-teams access)."""
    with get_conn() as conn:
        target = _load_target(conn, user_id, scope)
        team = _user_team(target["role"], body.team)
        if team == ALL_TEAMS and not scope.all_teams:
            raise HTTPException(403, "Only an All Teams admin can grant all-team access")
        # Their existing sessions carry no team, so nothing to invalidate — the
        # scope is read from the user row on every single request.
        conn.execute("UPDATE users SET team = ? WHERE id = ?", (team, user_id))
        conn.commit()
    log_activity(scope.user["name"], "user_team_changed",
                 f"{target['name']} → {TEAM_NAMES.get(team, team)}", team)
    return {"updated": True, "team": team}


class UserUpdate(BaseModel):
    """Every field optional — only what is sent is changed."""
    name: str | None = Field(default=None, min_length=2, max_length=80)
    role: str | None = Field(default=None, pattern="^(admin|sales)$")
    team: str | None = Field(default=None, max_length=20)
    can_upload: bool | None = None
    is_active: bool | None = None


@app.patch("/api/users/{user_id}")
def update_user(user_id: int, body: UserUpdate,
                scope: TeamScope = Depends(require_admin_scope)):
    """Edit a user: name, role, team, upload permission, active/disabled.

    Guards that matter more than the UI:
      * you cannot disable or demote yourself — that is how an installation
        ends up with nobody who can administer it;
      * the last active admin cannot be disabled or demoted for the same reason;
      * a role, team or active change bumps token_version, so the user's open
        sessions pick up the new rights (or lose them) on their next request.
    """
    admin = scope.user
    with get_conn() as conn:
        target = _load_target(conn, user_id, scope)

        role = body.role or target["role"]
        if body.role is not None or body.team is not None:
            team = _user_team(role, body.team if body.team is not None else target["team"])
        else:
            team = target["team"]
        if team == ALL_TEAMS and not scope.all_teams:
            raise HTTPException(403, "Only an All Teams admin can grant all-team access")
        if not scope.all_teams and team != scope.active:
            raise HTTPException(403, f"You can only move users within the {scope.active} team")

        active = target["is_active"] if body.is_active is None else body.is_active
        if user_id == admin["id"]:
            if not active:
                raise HTTPException(422, "You can't disable your own account")
            if role != "admin":
                raise HTTPException(422, "You can't remove your own admin role")
        if target["role"] == "admin" and (role != "admin" or not active) \
                and _active_admin_count(conn, exclude_id=user_id) == 0:
            raise HTTPException(422, "This is the last active admin — promote someone "
                                     "else first")

        can_upload = target["can_upload"] if body.can_upload is None else body.can_upload
        if role == "admin":
            can_upload = True          # admins always may add and edit displays
        name = (body.name or target["name"]).strip()

        rights_changed = (role != target["role"] or team != target["team"]
                          or bool(active) != bool(target["is_active"]))
        conn.execute(
            "UPDATE users SET name = ?, role = ?, team = ?, can_upload = ?, is_active = ?"
            + (", token_version = token_version + 1" if rights_changed else "")
            + " WHERE id = ?",
            (name, role, team, can_upload, active, user_id),
        )
        conn.commit()

    changes = []
    if name != target["name"]: changes.append(f"name → {name}")
    if role != target["role"]: changes.append(f"role → {role}")
    if team != target["team"]: changes.append(f"team → {TEAM_NAMES.get(team, team)}")
    if bool(can_upload) != bool(target["can_upload"]):
        changes.append("upload " + ("on" if can_upload else "off"))
    if bool(active) != bool(target["is_active"]):
        changes.append("enabled" if active else "disabled")
    log_activity(admin["name"], "user_updated",
                 f"{target['name']}: {', '.join(changes) or 'no change'}", team)
    log.info("User %s updated by %s: %s", target["username"], admin["name"], changes)
    return {"updated": True, "changes": changes}


class AdminPasswordReset(BaseModel):
    new_password: str = Field(min_length=6, max_length=120)


@app.post("/api/users/{user_id}/password")
def admin_reset_password(user_id: int, body: AdminPasswordReset,
                         scope: TeamScope = Depends(require_admin_scope)):
    """Set a new password for another user.

    Uses exactly the same hashing as a self-service change, and signs the user
    out of every device so an old session cannot outlive the reset. The password
    is never echoed back."""
    admin = scope.user
    with get_conn() as conn:
        target = _load_target(conn, user_id, scope)
        salt, ph = hash_pw(body.new_password)
        conn.execute(
            "UPDATE users SET salt = ?, pw_hash = ?, token_version = token_version + 1"
            " WHERE id = ?", (salt, ph, user_id),
        )
        conn.commit()
    log_activity(admin["name"], "password_reset", target["name"], target["team"])
    log.warning("Password reset for '%s' by %s", target["username"], admin["name"])
    return {"reset": True, "signed_out": True}


class PermissionIn(BaseModel):
    can_upload: bool


@app.patch("/api/users/{user_id}/permissions")
def set_permissions(user_id: int, body: PermissionIn,
                    scope: TeamScope = Depends(require_admin_scope)):
    admin = scope.user
    with get_conn() as conn:
        target = _load_target(conn, user_id, scope)
        conn.execute("UPDATE users SET can_upload = ? WHERE id = ?",
                     (body.can_upload, user_id))
        conn.commit()
    log_activity(admin["name"], "permission_changed",
                 f"{target['name']}: can_upload={'on' if body.can_upload else 'off'}",
                 target["team"])
    return {"updated": True}


@app.post("/api/users/{user_id}/logout-all")
def admin_logout_user(user_id: int, scope: TeamScope = Depends(require_admin_scope)):
    """Force-logout a user from every device."""
    admin = scope.user
    with get_conn() as conn:
        target = _load_target(conn, user_id, scope)
        conn.execute("UPDATE users SET token_version = token_version + 1 WHERE id = ?",
                     (user_id,))
        conn.commit()
    log_activity(admin["name"], "force_logout", target["name"], target["team"])
    return {"ok": True}


@app.get("/api/activity")
def activity(limit: int = 100, scope: TeamScope = Depends(require_admin_scope)):
    limit = max(1, min(limit, 500))
    where, params = scope.where("team")
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT to_char(at, 'DD Mon HH24:MI') AS at, user_name, action, detail, team"
            f" FROM activity_log WHERE {where} ORDER BY at DESC LIMIT ?",
            params + [limit],
        ).fetchall()
        return [dict(r) for r in rows]


@app.delete("/api/users/{user_id}")
def delete_user(user_id: int, scope: TeamScope = Depends(require_admin_scope)):
    admin = scope.user
    if user_id == admin["id"]:
        raise HTTPException(422, "You can't delete your own account")
    with get_conn() as conn:
        target = _load_target(conn, user_id, scope)
        if target["role"] == "admin" and _active_admin_count(conn, exclude_id=user_id) == 0:
            raise HTTPException(422, "This is the last active admin — promote someone "
                                     "else first")
        conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
        conn.commit()
    log_activity(admin["name"], "user_deleted", target["name"], target["team"])
    return {"deleted": True}


CAL_DAYS = 30  # availability horizon


# ---------- helpers ----------

def today() -> date:
    return date.today()


def slot_rows_for_screen(conn, screen_id: int):
    return conn.execute(
        """
        SELECT s.*, c.name AS campaign, c.creative, cl.company, cl.industry
        FROM slots s
        JOIN campaigns c ON c.id = s.campaign_id
        JOIN clients cl  ON cl.id = c.client_id
        WHERE s.screen_id = ?
        ORDER BY s.start_date
        """,
        (screen_id,),
    ).fetchall()


def occupancy(slots, on: date) -> list:
    """Slots covering a given day."""
    iso = on.isoformat()
    return [s for s in slots if s["start_date"] <= iso <= s["end_date"]]


def screen_payload(conn, screen) -> dict:
    slots = slot_rows_for_screen(conn, screen["id"])
    t = today()
    live = occupancy(slots, t)

    # Next opening: first day within horizon when occupied count drops
    next_opening = None
    for i in range(CAL_DAYS + 1):
        day = t + timedelta(days=i)
        if len(occupancy(slots, day)) < screen["loop_slots"]:
            next_opening = day.isoformat()
            break

    calendar = [
        {
            "date": (t + timedelta(days=i)).isoformat(),
            "booked": len(occupancy(slots, t + timedelta(days=i))),
        }
        for i in range(CAL_DAYS)
    ]

    return {
        **dict(screen),
        "live_count": len(live),
        "open_now": screen["loop_slots"] - len(live),
        "next_opening": next_opening,
        "live": [
            {
                "slot_id": s["id"],
                "company": s["company"],
                "campaign": s["campaign"],
                "creative": s["creative"],
                "position_no": s["position_no"],
                "start_date": s["start_date"],
                "end_date": s["end_date"],
                "days_left": (date.fromisoformat(s["end_date"]) - t).days,
            }
            for s in sorted(live, key=lambda r: r["position_no"])
        ],
        "calendar": calendar,
    }


# ---------- routes ----------

@app.get("/api/screens")
def list_screens(scope: TeamScope = Depends(team_scope)):
    """Full inventory with live occupancy, next opening and 30-day calendar —
    only the digital screens belonging to the caller's team."""
    where, params = scope.where("team")
    with get_conn() as conn:
        screens = conn.execute(
            f"SELECT * FROM screens WHERE {where} ORDER BY id", params).fetchall()
        return [screen_payload(conn, s) for s in screens]


@app.get("/api/screens/{screen_id}")
def get_screen(screen_id: int, scope: TeamScope = Depends(team_scope)):
    with get_conn() as conn:
        screen = conn.execute(
            "SELECT * FROM screens WHERE id = ?", (screen_id,)
        ).fetchone()
        scope.guard(screen, "Screen")
        return screen_payload(conn, screen)


@app.get("/api/campaigns/live")
def live_campaigns(scope: TeamScope = Depends(team_scope)):
    """Everything on air right now, across the team's screens."""
    iso = today().isoformat()
    where, params = scope.where("s.team")
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT s.id AS slot_id, s.position_no, s.start_date, s.end_date,
                   s.rate_month, s.booked_by, s.team, sc.name AS screen, sc.location,
                   c.name AS campaign, c.creative, cl.company, cl.industry
            FROM slots s
            JOIN screens sc  ON sc.id = s.screen_id
            JOIN campaigns c ON c.id = s.campaign_id
            JOIN clients cl  ON cl.id = c.client_id
            WHERE s.start_date <= ? AND s.end_date >= ? AND {where}
            ORDER BY s.end_date
            """.format(where=where),
            [iso, iso] + params,
        ).fetchall()
        return [
            {**dict(r), "days_left": (date.fromisoformat(r["end_date"]) - today()).days}
            for r in rows
        ]


@app.get("/api/clients")
def list_clients(scope: TeamScope = Depends(team_scope)):
    where, params = scope.where("team")
    with get_conn() as conn:
        return [dict(r) for r in conn.execute(
            f"SELECT * FROM clients WHERE {where} ORDER BY company", params)]


class Booking(BaseModel):
    screen_id: int = Field(gt=0)
    client_name: str = Field(min_length=2, max_length=100)
    campaign_name: str = Field(min_length=2, max_length=120)
    creative: str = Field(default="", max_length=200)
    start_date: date
    end_date: date
    rate_month: int = Field(default=0, ge=0)


@app.post("/api/campaigns", status_code=201)
def book_slot(b: Booking, scope: TeamScope = Depends(team_scope)):
    """Create a campaign + slot if the screen's loop has a free position
    for the entire requested window. The booking, its campaign and its client
    all inherit the screen's team, so a booking can never straddle the two."""
    user = scope.user
    if b.end_date < b.start_date:
        raise HTTPException(422, "end_date must be on or after start_date")
    with get_conn() as conn:
        screen = conn.execute(
            "SELECT * FROM screens WHERE id = ?", (b.screen_id,)
        ).fetchone()
        scope.guard(screen, "Screen")
        team = screen["team"]
        # Find-or-create the client by name within this team (case-insensitive).
        # The same company can be a client of both teams — those stay separate
        # records so neither team sees the other's pipeline.
        name = b.client_name.strip()
        row = conn.execute(
            "SELECT id FROM clients WHERE LOWER(company) = LOWER(?) AND team = ?",
            (name, team),
        ).fetchone()
        if row:
            client_id = row["id"]
        else:
            client_id = conn.execute(
                "INSERT INTO clients (company, team) VALUES (?,?)", (name, team)
            ).lastrowid
            log.info("New client auto-created: %s (%s)", name, team)

        # Positions blocked by any overlapping booking
        taken = {
            r["position_no"]
            for r in conn.execute(
                """SELECT position_no FROM slots
                   WHERE screen_id = ? AND NOT (end_date < ? OR start_date > ?)""",
                (b.screen_id, b.start_date.isoformat(), b.end_date.isoformat()),
            )
        }
        free = [p for p in range(1, screen["loop_slots"] + 1) if p not in taken]
        if not free:
            raise HTTPException(409, "No free loop position for that date range")

        cur = conn.cursor()
        cur.execute(
            "INSERT INTO campaigns (client_id, name, creative, team) VALUES (?,?,?,?)",
            (client_id, b.campaign_name, b.creative, team),
        )
        cur.execute(
            """INSERT INTO slots (screen_id, campaign_id, position_no,
               start_date, end_date, rate_month, booked_by, team)
               VALUES (?,?,?,?,?,?,?,?)""",
            (
                b.screen_id,
                cur.lastrowid,
                free[0],
                b.start_date.isoformat(),
                b.end_date.isoformat(),
                b.rate_month or round(
                    screen["rate_month"]
                    * ((b.end_date - b.start_date).days + 1) / 30
                ),
                user["name"],
                team,
            ),
        )
        conn.commit()
        log.info("Booked '%s' on screen %s (P%s) %s → %s [%s]",
                 b.campaign_name, b.screen_id, free[0], b.start_date, b.end_date, team)
        log_activity(user["name"], "booking_created",
                     f"{b.client_name} — {b.campaign_name} on screen {b.screen_id}",
                     team)
        return {"booked": True, "position_no": free[0], "campaign_id": cur.lastrowid}


class ScreenIn(BaseModel):
    name: str = Field(min_length=2, max_length=80)
    location: str = Field(min_length=2, max_length=120)
    city: str = Field(default="Bhubaneswar", max_length=60)
    width_ft: float = Field(gt=0, le=200)
    height_ft: float = Field(gt=0, le=200)
    res_w: int = Field(default=0, ge=0)
    res_h: int = Field(default=0, ge=0)
    loop_slots: int = Field(default=8, ge=1, le=24)
    slot_seconds: int = Field(default=15, ge=5, le=120)
    spots_per_day: int = Field(default=360, ge=1)
    rate_month: int = Field(default=0, ge=0)
    # Only read when the caller is viewing All Teams; a team user's screens
    # always land in their own team whatever the request body says.
    team: str | None = Field(default=None, max_length=20)


@app.post("/api/screens", status_code=201)
def add_screen(s: ScreenIn, scope: TeamScope = Depends(require_upload_scope)):
    """Add a new display to the network, inside the active team."""
    user = scope.user
    team = scope.write_team(s.team)
    with get_conn() as conn:
        cur = conn.execute(
            """INSERT INTO screens (name, location, city, width_ft, height_ft,
               res_w, res_h, loop_slots, slot_seconds, spots_per_day, rate_month, team)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (s.name, s.location, s.city, s.width_ft, s.height_ft, s.res_w,
             s.res_h, s.loop_slots, s.slot_seconds, s.spots_per_day, s.rate_month, team),
        )
        conn.commit()
        log.info("Screen added: %s (%s) [%s]", s.name, s.location, team)
        log_activity(user["name"], "screen_added", s.name, team)
        return {"created": True, "screen_id": cur.lastrowid, "team": team}


@app.put("/api/screens/{screen_id}")
def update_screen(screen_id: int, s: ScreenIn,
                  scope: TeamScope = Depends(require_upload_scope)):
    """Update an existing display's details. The team is not changed here —
    use PATCH /api/screens/{id}/team for that."""
    with get_conn() as conn:
        scope.guard(
            conn.execute("SELECT team FROM screens WHERE id = ?", (screen_id,)).fetchone(),
            "Screen")
        conn.execute(
            """UPDATE screens SET name=?, location=?, city=?, width_ft=?, height_ft=?,
               res_w=?, res_h=?, loop_slots=?, slot_seconds=?, spots_per_day=?, rate_month=?
               WHERE id=?""",
            (s.name, s.location, s.city, s.width_ft, s.height_ft, s.res_w, s.res_h,
             s.loop_slots, s.slot_seconds, s.spots_per_day, s.rate_month, screen_id),
        )
        conn.commit()
        log.info("Screen %s updated", screen_id)
        return {"updated": True}


class ScreenTeamIn(BaseModel):
    team: str = Field(max_length=20)


@app.patch("/api/screens/{screen_id}/team")
def move_screen_team(screen_id: int, body: ScreenTeamIn,
                     scope: TeamScope = Depends(require_admin_scope)):
    """Move a display — and everything booked on it — to the other team.
    All Teams admins only: it is the one operation that crosses the line."""
    if not scope.all_teams:
        raise HTTPException(403, "Switch to All Teams to move a display between teams")
    team = normalize_team(body.team)
    if team not in TEAMS:
        raise HTTPException(422, "Team must be Odisha or Raipur")
    with get_conn() as conn:
        screen = conn.execute(
            "SELECT name, team FROM screens WHERE id = ?", (screen_id,)).fetchone()
        if not screen:
            raise HTTPException(404, "Screen not found")
        conn.execute("UPDATE screens SET team = ? WHERE id = ?", (team, screen_id))
        # Bookings and their campaigns follow the screen, otherwise the receiving
        # team would own a display whose live ads it cannot see.
        conn.execute(
            "UPDATE campaigns SET team = ? WHERE id IN"
            " (SELECT campaign_id FROM slots WHERE screen_id = ?)", (team, screen_id))
        conn.execute("UPDATE slots SET team = ? WHERE screen_id = ?", (team, screen_id))
        conn.commit()
    log_activity(scope.user["name"], "screen_team_changed",
                 f"{screen['name']}: {screen['team']} → {team}", team)
    return {"updated": True, "team": team}


@app.delete("/api/screens/{screen_id}")
def delete_screen(screen_id: int, scope: TeamScope = Depends(require_upload_scope)):
    """Remove a display and all its bookings."""
    user = scope.user
    with get_conn() as conn:
        screen = scope.guard(
            conn.execute("SELECT team FROM screens WHERE id = ?", (screen_id,)).fetchone(),
            "Screen")
        conn.execute("DELETE FROM slots WHERE screen_id = ?", (screen_id,))
        conn.execute("DELETE FROM screens WHERE id = ?", (screen_id,))
        conn.commit()
        log.warning("Screen %s deleted with all bookings", screen_id)
        log_activity(user["name"], "screen_deleted", f"screen {screen_id}", screen["team"])
        return {"deleted": True}


@app.get("/api/bookings")
def all_bookings(scope: TeamScope = Depends(team_scope)):
    """Every booking (live, upcoming, recently ended) for the Manage view."""
    t = today()
    iso = t.isoformat()
    cutoff = (t - timedelta(days=14)).isoformat()
    where, params = scope.where("s.team")
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT s.id AS slot_id, s.position_no, s.start_date, s.end_date, s.rate_month,
                   s.booked_by, s.team, sc.name AS screen, c.name AS campaign, cl.company
            FROM slots s
            JOIN screens sc  ON sc.id = s.screen_id
            JOIN campaigns c ON c.id = s.campaign_id
            JOIN clients cl  ON cl.id = c.client_id
            WHERE s.end_date >= ? AND {where}
            ORDER BY s.start_date DESC
            """.format(where=where),
            [cutoff] + params,
        ).fetchall()
        out = []
        for r in rows:
            if r["start_date"] <= iso <= r["end_date"]:
                status = "live"
            elif r["start_date"] > iso:
                status = "upcoming"
            else:
                status = "ended"
            out.append({**dict(r), "status": status})
        return out


@app.patch("/api/slots/{slot_id}/stop")
def stop_slot(slot_id: int, scope: TeamScope = Depends(team_scope)):
    """Take an ad off air now: frees the loop position from today.
    Future bookings are removed entirely."""
    user = scope.user
    t = today().isoformat()
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM slots WHERE id = ?", (slot_id,)).fetchone()
        scope.guard(row, "Booking")
        if row["start_date"] > t:
            conn.execute("DELETE FROM slots WHERE id = ?", (slot_id,))
        else:
            yesterday = (today() - timedelta(days=1)).isoformat()
            conn.execute("UPDATE slots SET end_date = ? WHERE id = ?", (yesterday, slot_id))
        conn.commit()
        log.info("Slot %s taken off air", slot_id)
        log_activity(user["name"], "ad_stopped", f"slot {slot_id}", row["team"])
        return {"stopped": True}


@app.delete("/api/slots/{slot_id}")
def delete_slot(slot_id: int, scope: TeamScope = Depends(team_scope)):
    """Remove a booking entirely (wrong entry)."""
    with get_conn() as conn:
        row = conn.execute("SELECT team FROM slots WHERE id = ?", (slot_id,)).fetchone()
        scope.guard(row, "Booking")
        conn.execute("DELETE FROM slots WHERE id = ?", (slot_id,))
        conn.commit()
        log_activity(scope.user["name"], "booking_deleted", f"slot {slot_id}", row["team"])
        return {"deleted": True}


@app.delete("/api/clients/{client_id}")
def delete_client(client_id: int, scope: TeamScope = Depends(team_scope)):
    """Remove a client — cascades and removes all their bookings."""
    with get_conn() as conn:
        client = conn.execute(
            "SELECT company, team FROM clients WHERE id = ?", (client_id,)).fetchone()
        scope.guard(client, "Client")
        # Delete all slots (bookings) for this client's campaigns
        conn.execute(
            "DELETE FROM slots WHERE campaign_id IN "
            "(SELECT id FROM campaigns WHERE client_id = ?)",
            (client_id,),
        )
        # Delete campaigns
        conn.execute("DELETE FROM campaigns WHERE client_id = ?", (client_id,))
        # Delete client
        conn.execute("DELETE FROM clients WHERE id = ?", (client_id,))
        conn.commit()
        log.warning("Client %s deleted with all campaigns and bookings", client_id)
        log_activity(scope.user["name"], "client_deleted",
                     client["company"], client["team"])
        return {"deleted": True}


def bundled_photos_dir() -> Path:
    """The photos shipped with the repo (seed inventory)."""
    if getattr(sys, "frozen", False):
        return Path(sys._MEIPASS) / "frontend" / "photos"
    return Path(__file__).parent.parent / "frontend" / "photos"


def photos_dir() -> Path:
    """
    Writable photos folder.

    PHOTOS_DIR points at a persistent volume in production (a Render disk, a
    docker volume locally). Container filesystems are wiped on every deploy, so
    uploads must not live inside the image. Defaults to the original
    frontend/photos so local development is unchanged.
    """
    env = os.environ.get("PHOTOS_DIR", "").strip()
    if env:
        p = Path(env)
    elif getattr(sys, "frozen", False):
        p = Path(sys.executable).parent / "photos"
    else:
        p = Path(__file__).parent.parent / "frontend" / "photos"
    p.mkdir(parents=True, exist_ok=True)
    return p


def sync_seed_photos() -> None:
    """
    Copy the repo's seed photos into the writable photos folder on first boot.
    Uploaded files are never overwritten. Without this, the nine seeded screens
    would show broken images when PHOTOS_DIR points at an empty volume.
    """
    import shutil
    src_dir, dst_dir = bundled_photos_dir(), photos_dir()
    if not src_dir.exists() or src_dir.resolve() == dst_dir.resolve():
        return
    for f in src_dir.iterdir():
        if f.is_file() and not (dst_dir / f.name).exists():
            shutil.copy2(f, dst_dir / f.name)


@app.post("/api/screens/{screen_id}/photo")
async def upload_photo(screen_id: int, file: UploadFile = File(...),
                       scope: TeamScope = Depends(require_upload_scope)):
    """Attach a photo to a screen (jpg/png)."""
    ext = (file.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in ("jpg", "jpeg", "png", "webp"):
        raise HTTPException(422, "Use a jpg/png/webp image")
    data = await file.read()
    if len(data) > 8 * 1024 * 1024:
        raise HTTPException(413, "Image too large — keep it under 8 MB")
    with get_conn() as conn:
        scope.guard(
            conn.execute("SELECT team FROM screens WHERE id = ?", (screen_id,)).fetchone(),
            "Screen")
        fname = f"screen_{screen_id}.{ext}"
        (photos_dir() / fname).write_bytes(data)
        conn.execute("UPDATE screens SET photo = ? WHERE id = ?", (fname, screen_id))
        conn.commit()
    return {"photo": fname}


def _overlap_days(s: str, e: str, a: str, b: str) -> int:
    lo, hi = max(s, a), min(e, b)
    if lo > hi:
        return 0
    return (date.fromisoformat(hi) - date.fromisoformat(lo)).days + 1


def _all_slots(conn, scope: "TeamScope"):
    where, params = scope.where("s.team")
    return conn.execute(
        """SELECT s.*, sc.name AS screen, sc.loop_slots, c.name AS campaign, cl.company
           FROM slots s JOIN screens sc ON sc.id=s.screen_id
           JOIN campaigns c ON c.id=s.campaign_id JOIN clients cl ON cl.id=c.client_id
           WHERE {where}""".format(where=where),
        params,
    ).fetchall()


@app.get("/api/revenue")
def revenue(start: date | None = None, end: date | None = None,
            scope: TeamScope = Depends(require_admin_scope)):
    """Revenue = the actual Booking Amount of each booking, counted in full,
    attributed to the booking's start date. No proration, no demo values.
    Occupancy is the only day-based metric (it measures inventory, not money)."""
    t = today()
    end = end or t
    start = start or (end - timedelta(days=29))
    if start > end:
        raise HTTPException(422, "start must be on or before end")
    a, b = start.isoformat(), end.isoformat()

    where, params = scope.where("team")
    with get_conn() as conn:
        slots = _all_slots(conn, scope)
        screens = conn.execute(
            f"SELECT * FROM screens WHERE {where}", params).fetchall()
        clients = conn.execute(
            f"SELECT created_by FROM clients WHERE {where}", params).fetchall()
        campaign_count = conn.execute(
            f"SELECT COUNT(*) FROM campaigns WHERE {where}", params).fetchone()[0]

    def amt(s) -> int:
        return s["rate_month"] or 0   # stored as the booking's total amount

    def range_total(x: str, y: str) -> int:
        return sum(amt(s) for s in slots if x <= s["start_date"] <= y)

    iso = t.isoformat()
    month_a = t.replace(day=1).isoformat()
    year_a = t.replace(month=1, day=1).isoformat()

    by_screen, by_sales = {}, {}
    filtered_bookings = 0
    for s in slots:
        if not (a <= s["start_date"] <= b):
            continue
        filtered_bookings += 1
        r = amt(s)
        by_screen.setdefault(s["screen"], {"revenue": 0, "bookings": 0, "team": s["team"]})
        by_screen[s["screen"]]["revenue"] += r
        by_screen[s["screen"]]["bookings"] += 1
        who = s["booked_by"] or "Unassigned"
        # `teams` lets the All Teams view show which book a salesperson sold from;
        # in a single-team view it is always that one team.
        by_sales.setdefault(who, {"revenue": 0, "bookings": 0, "clients_added": 0,
                                  "teams": set()})
        by_sales[who]["revenue"] += r
        by_sales[who]["bookings"] += 1
        by_sales[who]["teams"].add(s["team"])
    for c in clients:
        who = c["created_by"]
        if who and who in by_sales:
            by_sales[who]["clients_added"] += 1
        elif who:
            by_sales.setdefault(who, {"revenue": 0, "bookings": 0, "clients_added": 1,
                                      "teams": set()})

    # occupancy over the filter window (inventory utilisation — day-based by nature)
    days_n = (end - start).days + 1
    cap_days = sum(sc["loop_slots"] for sc in screens) * days_n
    booked_days = sum(_overlap_days(s["start_date"], s["end_date"], a, b) for s in slots)
    occupancy = round(100 * booked_days / cap_days, 1) if cap_days else 0.0

    live_now = [s for s in slots if s["start_date"] <= iso <= s["end_date"]]
    total_loop = sum(sc["loop_slots"] for sc in screens)

    # 12-month trend: full booking amounts grouped by start month
    monthly = []
    y, m = t.year, t.month
    for _i in range(12):
        ma = date(y, m, 1)
        mb = date(y + (m == 12), (m % 12) + 1, 1) - timedelta(days=1)
        monthly.append({"month": ma.strftime("%b %y"),
                        "revenue": range_total(ma.isoformat(), mb.isoformat())})
        m -= 1
        if m == 0:
            y, m = y - 1, 12
    monthly.reverse()

    return {
        "range": {"start": a, "end": b},
        "team": scope.active,
        "team_name": TEAM_NAMES.get(scope.active, scope.active),
        "summary": {
            "filtered_revenue": range_total(a, b),
            "today": range_total(iso, iso),
            "month": range_total(month_a, iso),
            "year": range_total(year_a, iso),
            "bookings_in_range": filtered_bookings,
            "campaigns_total": campaign_count,
            "occupancy_pct": occupancy,
            "slots_booked_now": len(live_now),
            "slots_open_now": total_loop - len(live_now),
        },
        "by_screen": sorted(
            [{"screen": k, "revenue": v["revenue"], "bookings": v["bookings"],
              "team": v["team"]}
             for k, v in by_screen.items()], key=lambda x: -x["revenue"]),
        "by_sales": sorted(
            [{"name": k, "revenue": v["revenue"], "bookings": v["bookings"],
              "clients_added": v["clients_added"], "teams": sorted(v["teams"])}
             for k, v in by_sales.items()], key=lambda x: -x["revenue"]),
        "monthly": monthly,
    }


@app.get("/api/export/bookings.csv")
def export_csv(start: date | None = None, end: date | None = None,
               scope: TeamScope = Depends(require_admin_scope)):
    import csv
    import io
    t = today()
    end = end or t
    start = start or (end - timedelta(days=29))
    a, b = start.isoformat(), end.isoformat()
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Team", "Screen", "Client", "Campaign", "Position", "Start", "End",
                "Booking Amount", "Booked By"])
    with get_conn() as conn:
        for s in _all_slots(conn, scope):
            if not (a <= s["start_date"] <= b):
                continue
            w.writerow([TEAM_NAMES.get(s["team"], s["team"]),
                        s["screen"], s["company"], s["campaign"], s["position_no"],
                        s["start_date"], s["end_date"], s["rate_month"] or 0,
                        s["booked_by"] or ""])
    buf.seek(0)
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="bookings_{a}_to_{b}.csv"'})


@app.get("/api/export/revenue.xlsx")
def export_xlsx(start: date | None = None, end: date | None = None,
                scope: TeamScope = Depends(require_admin_scope)):
    import io
    from openpyxl import Workbook
    data = revenue(start, end, scope)
    wb = Workbook()
    ws = wb.active; ws.title = "Summary"
    ws.append([f"MediaTrack Revenue Report — {data['team_name']}",
               f"{data['range']['start']} to {data['range']['end']}"])
    for k, v in data["summary"].items():
        ws.append([k.replace("_", " ").title(), v])
    ws2 = wb.create_sheet("By Screen")
    ws2.append(["Screen", "Revenue", "Bookings"])
    for r in data["by_screen"]:
        ws2.append([r["screen"], r["revenue"], r["bookings"]])
    ws3 = wb.create_sheet("By Salesperson")
    ws3.append(["Salesperson", "Revenue", "Bookings", "Clients Added"])
    for r in data["by_sales"]:
        ws3.append([r["name"], r["revenue"], r["bookings"], r["clients_added"]])
    ws4 = wb.create_sheet("Monthly Trend")
    ws4.append(["Month", "Revenue"])
    for r in data["monthly"]:
        ws4.append([r["month"], r["revenue"]])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="revenue_report.xlsx"'})


@app.get("/api/export/availability.xlsx")
def export_availability(scope: TeamScope = Depends(team_scope)):
    """One click → polished Excel of the team's screen availability, for any
    logged-in user (sales send this to clients)."""
    user = scope.user
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    t = today()
    where, params = scope.where("team")
    with get_conn() as conn:
        screens = conn.execute(
            f"SELECT * FROM screens WHERE {where} ORDER BY name", params).fetchall()
        payloads = [screen_payload(conn, s) for s in screens]

    wb = Workbook()
    HEAD = Font(bold=True, color="FFFFFF", size=11)
    FILL = PatternFill("solid", fgColor="C11527")
    THIN = Border(*[Side(style="thin", color="DDDDDD")] * 4)
    GREEN = Font(color="1A7A44", bold=True)
    RED = Font(color="C11527", bold=True)

    ws = wb.active
    ws.title = "Available Media"
    ws.append([f"AJANTA ADVERTISERS — {TEAM_NAMES.get(scope.active, scope.active)} —"
               f" Digital Media Availability — {t.strftime('%d %b %Y')}"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    cols = ["Screen", "Location", "City", "Size (ft)", "Resolution",
            "Loop Slots", "Booked Now", "Open Now", "Next Opening",
            "Card Rate (₹/mo)", "Status"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=2, column=c)
        cell.font = HEAD
        cell.fill = FILL
        cell.alignment = Alignment(horizontal="center")
    for p in sorted(payloads, key=lambda x: -x["open_now"]):
        status = "AVAILABLE" if p["open_now"] else "FULL"
        ws.append([
            p["name"], p["location"], p["city"],
            f"{p['width_ft']}x{p['height_ft']}",
            f"{p['res_w']}x{p['res_h']}",
            p["loop_slots"], p["live_count"], p["open_now"],
            p["next_opening"] or "—", p["rate_month"], status,
        ])
        r = ws.max_row
        ws.cell(row=r, column=8).font = GREEN if p["open_now"] else RED
        ws.cell(row=r, column=11).font = GREEN if p["open_now"] else RED
        for c in range(1, len(cols) + 1):
            ws.cell(row=r, column=c).border = THIN
    widths = [22, 30, 14, 10, 12, 10, 11, 10, 13, 15, 11]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

    ws2 = wb.create_sheet("30-Day Grid")
    dates = [(t + timedelta(days=i)) for i in range(30)]
    ws2.append(["Screen"] + [d.strftime("%d %b") for d in dates])
    for c in range(1, 32):
        cell = ws2.cell(row=1, column=c)
        cell.font = HEAD
        cell.fill = FILL
    ok_fill = PatternFill("solid", fgColor="E8F5EE")
    full_fill = PatternFill("solid", fgColor="FBE3E5")
    for p in payloads:
        row = [p["name"]] + [p["loop_slots"] - d["booked"] for d in p["calendar"]]
        ws2.append(row)
        r = ws2.max_row
        for c in range(2, 32):
            cell = ws2.cell(row=r, column=c)
            cell.fill = ok_fill if cell.value else full_fill
            cell.alignment = Alignment(horizontal="center")
    ws2.column_dimensions["A"].width = 22
    for i in range(2, 32):
        ws2.column_dimensions[get_column_letter(i)].width = 7
    ws2.freeze_panes = "B2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log_activity(user["name"], "availability_exported", "", scope.active)
    fname = f"Ajanta_Available_Media_{scope.active}_{t.isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/api/backup")
def download_backup(scope: TeamScope = Depends(require_admin_scope)):
    """
    Download a full backup of the live database.

    There is no database file to copy any more, so this streams a pg_dump —
    restore it anywhere with:  psql "$DATABASE_URL" < mediatrack_backup_....sql
    If pg_dump is unavailable, every table is exported as JSON instead, so the
    button never dead-ends.
    """
    import io
    import subprocess
    stamp = date.today().isoformat()
    if not scope.all_teams:
        # A single-team view must not receive a full-cluster dump; fall through
        # to the JSON export, which is filtered per team.
        return _json_backup(scope, stamp)
    try:
        out = subprocess.run(
            ["pg_dump", "--no-owner", "--no-privileges", database_url()],
            capture_output=True, timeout=120,
        )
        if out.returncode == 0 and out.stdout:
            return StreamingResponse(
                iter([out.stdout]), media_type="application/sql",
                headers={"Content-Disposition":
                         f'attachment; filename="mediatrack_backup_{stamp}.sql"'},
            )
        log.warning("pg_dump failed (%s) — falling back to JSON backup",
                    out.stderr.decode()[:200])
    except (FileNotFoundError, subprocess.TimeoutExpired) as e:
        log.warning("pg_dump unavailable (%s) — falling back to JSON backup", e)

    return _json_backup(scope, stamp)


def _json_backup(scope: "TeamScope", stamp: str):
    """Table-by-table JSON export, scoped to the teams the caller may see."""
    import io
    import json
    where, params = scope.where("team")
    dump = {}
    with get_conn() as conn:
        for table in ("clients", "screens", "campaigns", "slots", "users"):
            dump[table] = [
                dict(r) for r in conn.execute(
                    f"SELECT * FROM {table} WHERE {where} ORDER BY id", params)
            ]
    suffix = "" if scope.all_teams else f"_{scope.active}"
    buf = io.BytesIO(json.dumps(dump, indent=2, default=str).encode())
    return StreamingResponse(
        buf, media_type="application/json",
        headers={"Content-Disposition":
                 f'attachment; filename="mediatrack_backup{suffix}_{stamp}.json"'},
    )


APP_VERSION = "4.0"


@app.get("/api/version")
def version():
    return {"version": APP_VERSION}


@app.get("/api/health")
def health(response: Response):
    """
    Liveness + readiness for Render's health check.

    Unauthenticated on purpose (the platform probe has no session) and it
    reports nothing sensitive: an app that answers 200 but cannot reach its
    database is not actually healthy, so the DB round-trip is the point.
    """
    checks = {"database": "unknown", "photos": "unknown"}
    ok = True
    try:
        with get_conn() as conn:
            conn.execute("SELECT 1").fetchone()
        checks["database"] = "ok"
    except Exception as e:                                   # noqa: BLE001
        log.exception("Health check: database unreachable")
        checks["database"] = "unreachable"
        ok = False
    try:
        checks["photos"] = "ok" if os.access(photos_dir(), os.W_OK) else "not writable"
        ok = ok and checks["photos"] == "ok"
    except Exception:                                        # noqa: BLE001
        checks["photos"] = "unavailable"
        ok = False

    response.status_code = 200 if ok else 503
    return {"status": "healthy" if ok else "unhealthy",
            "version": APP_VERSION, "environment": ENV, "checks": checks}


@app.get("/")
def index_page():
    """Serve the dashboard with no-store so browsers never cache a stale UI."""
    f = _find_frontend() / "index.html"
    if not f.exists():
        raise HTTPException(404, "UI not found")
    return FileResponse(f, headers={"Cache-Control": "no-store, must-revalidate"})


def _find_frontend() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys._MEIPASS) / "frontend"
    return Path(__file__).parent.parent / "frontend"


# ---------- serve the dashboard itself ----------
# Anyone on the office network can open http://<this-pc-ip>:8000 — no file sharing needed.
if getattr(sys, "frozen", False):
    _frontend = Path(sys._MEIPASS) / "frontend"  # bundled inside the .exe
else:
    _frontend = Path(__file__).parent.parent / "frontend"
sync_seed_photos()
app.mount("/photos", StaticFiles(directory=photos_dir()), name="photos")
if _frontend.exists():
    app.mount("/", StaticFiles(directory=_frontend, html=True), name="ui")
